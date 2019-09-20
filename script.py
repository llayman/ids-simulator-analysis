from datetime import datetime
from openpyxl import Workbook
import os
from pathlib import Path
import psycopg2
import shlex
import shutil
from sqlalchemy import create_engine, desc
from sqlalchemy.orm import sessionmaker
from subprocess import Popen

# Constants that may need to be changed based on local machine configuration
HEROKU_APP = 'cry-wolf'
SNAPSHOTS_DIR = 'snapshots'
PG_USERNAME = 'postgres'
PG_PASSWORD = 'postgres'
PG_HOST = 'localhost'
PG_DATABASE = 'crywolf'

EXCEL_DIR = 'excel'


def download_and_import():
    # Download the database dump
    Popen(shlex.split(f"heroku pg:backups:capture -a {HEROKU_APP}")).wait()
    Popen(shlex.split(f"heroku pg:backups:download -a {HEROKU_APP}")).wait()

    # copy it to snapshots folder and timestamp it
    if not os.path.exists(SNAPSHOTS_DIR):
        os.makedirs(SNAPSHOTS_DIR)
    snapshot = Path(SNAPSHOTS_DIR, f'{HEROKU_APP}_{datetime.now().strftime("%Y%m%d_%H-%M-%S")}.dump')
    shutil.copy('latest.dump', snapshot)
    os.remove('latest.dump')

    # Need to set this so Postgres can log in. Not recommended for security reasons on a server.
    # Assumes Postgres is running on localhost
    os.environ["PGPASSWORD"] = PG_PASSWORD

    # Do the import using the pg_restore tool
    Popen(shlex.split(f"pg_restore --verbose --clean --no-acl --no-owner -h {PG_HOST} -U {PG_USERNAME} -d {PG_DATABASE} {snapshot}")).wait()

    print(f"Run the following from .venv terminal: sqlacodegen postgresql:///{PG_DATABASE} --outfile models.py")


# Convert a Sqlaclhemy model with data to a dictionary.
def _to_dict(model):
    if model is None:
        return {}
    d = model.__dict__
    d.pop('_sa_instance_state', None)   # remove this sqlalchemy state variable
    return d

def compute_results(session, decided_only=False):
    events = {e.id: _to_dict(e) for e in session.query(Event)}
    users = {u.username: _to_dict(u) for u in session.query(User)}

    results = []

    for username, user in users.items():
        user_events = user['events'].split(',')

        # Get all of the decisions made for events by a user
        decisions = [_to_dict(d) for d in session.query(EventDecision).filter_by(user=username)] 
        latest_decisions = {}
        decision_counts = {}
        i_dont_knows = 0

        for d in decisions:
            # Increment how many times a decision was made for a particular event
            decision_counts[d["event_id"]] = decision_counts.get(d["event_id"], 0) + 1

            # TODO: Do something with "Check" events - ids 74, 75
            d['TP'] = False
            d['FP'] = False
            d['TN'] = False
            d['FN'] = False
            
            # "should_escalate" = 1 for TP, 0 for TN
            # Determine whether a decision is correct and a TP, FP, TN, or FN
            should_escalate = True if events[d['event_id']]['should_escalate'] == '1' else False
            if d['escalate'] == "I don't know":
                d['correct?'] = False
                i_dont_knows += 1
            elif should_escalate and d['escalate'] == "Escalate":
                d['TP'] = True
                d['state'] = 'TP'
                d['correct?'] = True
            elif should_escalate and d['escalate'] == "Don't escalate":
                d['FN'] = True
                d['state'] = 'FN'
                d['correct?'] = False
            elif not should_escalate and d['escalate'] == "Don't escalate":
                d['TN'] = True
                d['state'] = 'TN'
                d['correct?'] = True
            elif not should_escalate and d['escalate'] == 'Escalate':
                d['FP'] = True
                d['state'] = 'FP'
                d['correct?'] = False
            else:
                raise Exception("Encountered an unknown value for 'Escalate' in the event decision", d['escalate'])

            # Update the "most recent decision" for each event
            if d["event_id"] not in latest_decisions or d["time_event_decision"] > latest_decisions[d["event_id"]]["time_event_decision"]:
                # print("new latest:",
                #     d["user"],
                #     d["event_id"],
                #     'None' if d["event_id"] not in latest_decisions else latest_decisions[d["event_id"]]["time_event_decision"],
                #     "-->",
                #     d["time_event_decision"])

                latest_decisions[d["event_id"]] = d

        confusion = {
            'TP' : 0,
            'FP' : 0,
            'TN' : 0,
            'FN' : 0
        }
        num_correct = 0
        num_with_confidence = 0
        confidence_sum = 0
        # Look at most recent decisions to compute correctness, confidence, and confusion matrix
        for decision in latest_decisions.values():
            if 'confidence' in decision and decision['confidence'] != 'None':
                confidence_sum += int(decision['confidence'])
                num_with_confidence += 1
            if decision['correct?']:
                num_correct += 1
            # Generate confusion matrix for event decisions made. "I don't knows" are excluded
            if 'state' in decision:
                confusion[decision['state']] += 1
        
        specificity = 0 if confusion['TN'] + confusion['FP'] == 0 else confusion['TN'] / (confusion['TN'] + confusion['FP'])
        sensitivity = 0 if confusion['TP'] + confusion['FN'] == 0 else confusion['TP'] / (confusion['TP'] + confusion['FN'])
        precision = 0 if confusion['TP'] + confusion['FP'] == 0 else confusion['TP'] / (confusion['TP'] + confusion['FP'])

        if len(latest_decisions):
            print(f"{username} - " 
                f"{len(latest_decisions)}/{(len(user_events) + 3)} decided, "     # the 3 are the 2 check events + 1 obvious attack everyone got
                f"{confidence_sum / num_with_confidence if latest_decisions else 0:0.1f} avg confidence, "
                f"{num_correct * 100 / len(latest_decisions) if latest_decisions else 0:.0f}% correct, " 
                f"{i_dont_knows} IDKs, "
                f"{specificity * 100:.1f} specificity, "
                f"{sensitivity * 100:.1f} sensitivity, "
                f"{precision * 100:.1f} precision, "
                f"{confusion}"
                )
        
        # Add computed values into user dictionary
        user['decided'] = len(latest_decisions)
        user['perc_decided'] = len(latest_decisions) * 100 / (len(user_events) + 3)
        user['avg_confidence'] = confidence_sum / len(latest_decisions) if latest_decisions else 0
        user['correct'] = num_correct
        user['perc_correct'] = num_correct * 100 / len(latest_decisions) if latest_decisions else 0
        user['i_dont_knows'] = i_dont_knows
        user['sensitivity'] = sensitivity
        user['specificity'] = specificity
        user['precision'] = precision
        user['TP'] = confusion['TP']
        user['FP'] = confusion['FP']
        user['TN'] = confusion['TN']
        user['FN'] = confusion['FN']
        
        # get prequestionnaire answers for user and rename dictionary keys with prefix 'preq_'
        preq = {}
        for k, v in _to_dict(session.query(PrequestionnaireAnswer).filter_by(user=username).order_by(desc(PrequestionnaireAnswer.timestamp)).first()).items():
            preq[f'preq_{k}'] = v

        # get survey answers for user and rename dictionary keys with prefix 'surv_'
        survey = {}
        for k, v in _to_dict(session.query(SurveyAnswer).filter_by(user=username).order_by(desc(SurveyAnswer.timestamp)).first()).items():
            survey[f'surv_{k}'] = v
        
        # create 'master' user data with user, computations, prequestionnaire answers, and survey answers
        if not decided_only or user['decided'] > 0:
            results.append({**user, **preq, **survey})  
    
    return results
    
def write_excel(results, session):
    if not os.path.exists(EXCEL_DIR):
        os.makedirs(EXCEL_DIR)
    excel_file = Path(EXCEL_DIR, f'{HEROKU_APP}_{datetime.now().strftime("%Y%m%d_%H-%M-%S")}.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = 'master'

    headers = \
        ['id', 
        'username', 
        'group',
        'time_begin', 
        'time_end',
        'questionnaire_complete', 
        'training_complete',
        'survey_complete', 
        'decided',
        'perc_decided',
        'avg_confidence',
        'correct',
        'perc_correct',
        'i_dont_knows',
        'TP',
        'FP',
        'TN',
        'FN',
        'preq_timestamp',
        'preq_role',
        'preq_exp_researcher',
        'preq_exp_admin',
        'preq_exp_software',
        'preq_exp_security',
        'preq_exp_security',
        'preq_familiarity_none',
        'preq_familiarity_read',
        'preq_familiarity_controlled',
        'preq_familiarity_public',
        'preq_familiarity_engineered',
        'preq_subnet_mask',
        'preq_network_address',
        'preq_tcp_faster',
        'preq_http_port',
        'preq_firewall',
        'preq_socket',
        'preq_which_model',
        'surv_timestamp',
        'surv_mental',
        'surv_physical',
        'surv_temporal',
        'surv_performance',
        'surv_effort',
        'surv_frustration',
        'surv_useful_info',
        'surv_feedback'
        ]

    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])

    for row in range(2, len(results)+2):
        for field in range(len(headers)):
            ws.cell(row=row, column=field+1, value=results[row-2].get(headers[field], None))

    _create_sheet_for_table(wb, 'User', User)
    _create_sheet_for_table(wb, 'PrequestionnaireAnswer', PrequestionnaireAnswer)
    _create_sheet_for_table(wb, 'TrainingEvent', TrainingEvent)
    _create_sheet_for_table(wb, 'TrainingEventDecision', TrainingEventDecision)
    _create_sheet_for_table(wb, 'Event', Event)
    _create_sheet_for_table(wb, 'EventClicked', EventClicked)
    _create_sheet_for_table(wb, 'EventDecision', EventDecision)
    _create_sheet_for_table(wb, 'SurveyAnswer', SurveyAnswer)

    wb.save(excel_file)

def _create_sheet_for_table(wb, sheet_name, model):
    ws = wb.create_sheet(sheet_name)
    results = [_to_dict(d) for d in session.query(model)]
    headers = list(results[0].keys())

    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])
    for row in range(2, len(results) + 2):
        for field in range(len(headers)):
            ws.cell(row=row, column=field+1, value=results[row-2].get(headers[field], None))


if __name__ == "__main__":

    # 0. Must have run 'heroku login' from prior to running this script

    # 1. download_and_import first. Must manually generate models after that.
    download_and_import()

    # 2. Manually generate models using sqlacodegen string from 1.

    # 3. compute_results
    engine = create_engine(f'postgresql+psycopg2://{PG_USERNAME}:{PG_PASSWORD}@{PG_HOST}/{PG_DATABASE}')
    Session = sessionmaker(bind=engine)
    session = Session()
    from models import User, Event, EventClicked, EventDecision, PrequestionnaireAnswer, TrainingEvent, TrainingEventDecision, SurveyAnswer
    results = compute_results(session, decided_only=True)

    # 4. write_excel
    write_excel(results, session)

    session.close()